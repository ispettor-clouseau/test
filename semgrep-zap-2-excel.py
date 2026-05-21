#!/usr/bin/env python3
"""
semgrep2excel.py — SAST + DAST → Excel/CSV
Combina finding Semgrep e alert OWASP ZAP in un unico report Excel.
Compatibile con Windows e macOS | Python 3.8+

USO:
  # Solo Semgrep
  python3 semgrep2excel.py --json semgrep.json

  # Solo ZAP
  python3 semgrep2excel.py --zap zap_report.json
  python3 semgrep2excel.py --zap zap_report.xml

  # Semgrep + ZAP insieme (report unificato)
  python3 semgrep2excel.py --json semgrep.json --zap zap_report.json

  # Output personalizzato
  python3 semgrep2excel.py --json semgrep.json --zap zap_report.xml -o report_finale.xlsx

  # Target diretto (lancia semgrep internamente)
  python3 semgrep2excel.py UserService.java --zap zap_report.json

  # CSV
  python3 semgrep2excel.py --json semgrep.json --zap zap_report.json --csv

INSTALLAZIONE DIPENDENZE (una sola volta):
  pip install openpyxl
"""

import json
import sys
import csv
import os
import subprocess
import shutil
import argparse
import xml.etree.ElementTree as ET
from datetime import datetime

# ---------------------------------------------------------------------------
# MAPPE DI RIFERIMENTO
# ---------------------------------------------------------------------------

SEVERITY_ORDER = {"CRITICO": 1, "ALTO": 2, "MEDIO": 3, "BASSO": 4, "INFO": 5}

CWE_LABELS = {
    "CWE-89":  "SQL Injection",
    "CWE-79":  "Cross-Site Scripting (XSS)",
    "CWE-78":  "Command Injection",
    "CWE-22":  "Path Traversal",
    "CWE-798": "Credenziali Hardcoded",
    "CWE-614": "Cookie non sicuro",
    "CWE-209": "Esposizione Stack Trace",
    "CWE-502": "Deserializzazione non sicura",
    "CWE-918": "SSRF",
    "CWE-347": "JWT - Algoritmo non verificato",
    "CWE-532": "Dati sensibili nei log",
    "CWE-312": "Secret in testo chiaro",
    "CWE-16":  "Configurazione errata",
    "CWE-693": "Header di sicurezza mancante",
    "CWE-319": "Trasmissione in chiaro",
    "CWE-352": "CSRF",
    "CWE-200": "Esposizione dati sensibili",
}

# ZAP riskcode → severità italiana
ZAP_RISK = {"3": "CRITICO", "2": "ALTO", "1": "MEDIO", "0": "BASSO"}

SEVERITY_IT = {"ERROR": "CRITICO", "WARNING": "ALTO", "INFO": "MEDIO"}

# Colori Excel per severità
COLORS = {
    "CRITICO": "FCE4D6",
    "ALTO":    "FFF2CC",
    "MEDIO":   "EBF3FB",
    "BASSO":   "E2EFDA",
    "INFO":    "F2F2F2",
}

HEADER_COLOR = "1F4E79"


# ---------------------------------------------------------------------------
# PARSING SEMGREP
# ---------------------------------------------------------------------------

def parse_semgrep(data):
    rows = []
    for f in data.get("results", []):
        meta = f["extra"].get("metadata", {})

        def to_str(val):
            if isinstance(val, list):
                return ", ".join(str(v) for v in val) if val else ""
            return str(val) if val else ""

        cwe_raw  = to_str(meta.get("cwe", ""))
        cwe_id   = cwe_raw.split(",")[0].strip()
        cwe_id   = cwe_id if cwe_id.startswith("CWE-") else ""
        cwe_desc = CWE_LABELS.get(cwe_id, "")
        owasp    = to_str(meta.get("owasp", ""))
        sev_raw  = f["extra"].get("severity", "INFO")
        sev      = SEVERITY_IT.get(sev_raw, sev_raw)

        rows.append({
            "Fonte":          "SAST — Semgrep",
            "File / URL":     f.get("path", ""),
            "Riga / Param":   str(f["start"]["line"]),
            "Regola / Alert": f.get("check_id", ""),
            "Severità":       sev,
            "CWE":            cwe_id,
            "Descrizione CWE":cwe_desc,
            "OWASP":          owasp,
            "Messaggio":      f["extra"].get("message", ""),
            "Soluzione":      "",
            "Stato fix":      "Aperto",
            "Note":           "",
        })
    return rows


# ---------------------------------------------------------------------------
# PARSING ZAP JSON
# ---------------------------------------------------------------------------

def parse_zap_json(data):
    rows = []

    # ZAP JSON può avere struttura {"site": [...]} o {"@version":..., "site":[...]}
    sites = data if isinstance(data, list) else data.get("site", [])
    if isinstance(sites, dict):
        sites = [sites]

    for site in sites:
        host    = site.get("@host", site.get("host", ""))
        alerts  = site.get("alerts", [])

        for alert in alerts:
            riskcode = str(alert.get("riskcode", alert.get("@riskcode", "0")))
            sev      = ZAP_RISK.get(riskcode, "INFO")

            cweid    = str(alert.get("cweid", alert.get("@cweid", "")))
            cwe_id   = f"CWE-{cweid}" if cweid and cweid != "0" else ""
            cwe_desc = CWE_LABELS.get(cwe_id, "")

            name     = alert.get("name",     alert.get("alert", ""))
            desc     = alert.get("desc",     "")
            solution = alert.get("solution", "")

            # Istanze (URL + parametro)
            instances = alert.get("instances", [])
            if not instances:
                # alert senza istanze → riga unica con host
                rows.append(_zap_row(host, "", name, sev, cwe_id, cwe_desc,
                                     desc, solution))
            else:
                for inst in instances:
                    uri    = inst.get("uri",   inst.get("url", host))
                    param  = inst.get("param", inst.get("parameter", ""))
                    rows.append(_zap_row(uri, param, name, sev, cwe_id,
                                        cwe_desc, desc, solution))
    return rows


def _zap_row(url, param, name, sev, cwe_id, cwe_desc, desc, solution):
    # Tronca desc a 300 char per leggibilità in Excel
    desc_short = desc.replace("<p>", "").replace("</p>", " ").strip()
    desc_short = desc_short[:300] + "…" if len(desc_short) > 300 else desc_short
    sol_short  = solution.replace("<p>", "").replace("</p>", " ").strip()
    sol_short  = sol_short[:200] + "…" if len(sol_short) > 200 else sol_short
    return {
        "Fonte":          "DAST — ZAP",
        "File / URL":     url,
        "Riga / Param":   param,
        "Regola / Alert": name,
        "Severità":       sev,
        "CWE":            cwe_id,
        "Descrizione CWE":cwe_desc,
        "OWASP":          "",
        "Messaggio":      desc_short,
        "Soluzione":      sol_short,
        "Stato fix":      "Aperto",
        "Note":           "",
    }


# ---------------------------------------------------------------------------
# PARSING ZAP XML
# ---------------------------------------------------------------------------

def parse_zap_xml(path):
    rows = []
    try:
        tree = ET.parse(path)
        root = tree.getroot()
    except ET.ParseError as e:
        print(f"❌ XML ZAP non valido: {e}")
        sys.exit(1)

    # Struttura: <OWASPZAPReport> → <site> → <alerts> → <alertitem>
    for site in root.findall(".//site"):
        for alert in site.findall(".//alertitem"):
            def txt(tag):
                el = alert.find(tag)
                return el.text.strip() if el is not None and el.text else ""

            riskcode = txt("riskcode")
            sev      = ZAP_RISK.get(riskcode, "INFO")
            cweid    = txt("cweid")
            cwe_id   = f"CWE-{cweid}" if cweid and cweid != "0" else ""
            cwe_desc = CWE_LABELS.get(cwe_id, "")
            name     = txt("alert") or txt("name")
            desc     = txt("desc")
            solution = txt("solution")

            instances = alert.findall(".//instance")
            if not instances:
                rows.append(_zap_row("", "", name, sev, cwe_id, cwe_desc,
                                     desc, solution))
            else:
                for inst in instances:
                    uri   = inst.find("uri")
                    param = inst.find("param")
                    rows.append(_zap_row(
                        uri.text.strip()   if uri   is not None and uri.text   else "",
                        param.text.strip() if param is not None and param.text else "",
                        name, sev, cwe_id, cwe_desc, desc, solution
                    ))
    return rows


# ---------------------------------------------------------------------------
# CARICAMENTO FILE ZAP (auto-detect JSON / XML)
# ---------------------------------------------------------------------------

def load_zap(path):
    if not os.path.isfile(path):
        print(f"❌ File ZAP non trovato: {path}")
        sys.exit(1)

    ext = os.path.splitext(path)[1].lower()

    # Auto-detect: prova JSON, poi XML
    if ext == ".xml":
        print(f"📂 Lettura ZAP XML: {path}")
        return parse_zap_xml(path)

    print(f"📂 Lettura ZAP JSON: {path}")
    with open(path, encoding="utf-8") as f:
        try:
            data = json.load(f)
            return parse_zap_json(data)
        except json.JSONDecodeError:
            # Potrebbe essere XML con estensione sbagliata
            print("   (non è JSON valido, provo come XML…)")
            return parse_zap_xml(path)


# ---------------------------------------------------------------------------
# ORDINAMENTO
# ---------------------------------------------------------------------------

def sort_rows(rows):
    return sorted(rows, key=lambda r: (
        SEVERITY_ORDER.get(r["Severità"], 9),
        r["Fonte"],
        r["File / URL"],
    ))


# ---------------------------------------------------------------------------
# OUTPUT CSV
# ---------------------------------------------------------------------------

def to_csv(rows, output_path):
    if not rows:
        print("⚠  Nessun finding. CSV non generato.")
        return
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    print(f"✅ CSV salvato: {output_path}  ({len(rows)} finding)")


# ---------------------------------------------------------------------------
# OUTPUT EXCEL
# ---------------------------------------------------------------------------

def to_excel(rows, output_path, semgrep_rows=None, zap_rows=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ openpyxl non trovato. Installa con: pip install openpyxl")
        sys.exit(1)

    if not rows:
        print("⚠  Nessun finding. Excel non generato.")
        return

    wb  = Workbook()
    thin = Side(style="thin", color="D0D0D0")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def make_sheet(ws, sheet_rows, title):
        ws.title = title
        headers  = list(sheet_rows[0].keys()) if sheet_rows else list(rows[0].keys())

        # Header row
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill      = PatternFill("solid", fgColor=HEADER_COLOR)
            c.font      = Font(color="FFFFFF", bold=True, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border    = brd
        ws.row_dimensions[1].height = 26

        # Data rows
        for ri, row in enumerate(sheet_rows, 2):
            sev  = row.get("Severità", "INFO")
            fill = PatternFill("solid", fgColor=COLORS.get(sev, "FFFFFF"))
            for ci, key in enumerate(headers, 1):
                c = ws.cell(row=ri, column=ci, value=row.get(key, ""))
                c.fill      = fill
                c.font      = Font(size=10)
                c.border    = brd
                c.alignment = Alignment(vertical="top", wrap_text=True)

        # Larghezze colonne
        widths = {
            "Fonte": 16, "File / URL": 36, "Riga / Param": 12,
            "Regola / Alert": 28, "Severità": 11, "CWE": 10,
            "Descrizione CWE": 24, "OWASP": 14, "Messaggio": 46,
            "Soluzione": 32, "Stato fix": 13, "Note": 20,
        }
        for ci, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 16)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes    = "A2"

    # --- Foglio principale: tutti i finding ---
    ws_all = wb.active
    make_sheet(ws_all, rows, "Tutti i Finding")

    # --- Foglio SAST (se presenti) ---
    if semgrep_rows:
        ws_sast = wb.create_sheet("SAST — Semgrep")
        make_sheet(ws_sast, semgrep_rows, "SAST — Semgrep")

    # --- Foglio DAST (se presenti) ---
    if zap_rows:
        ws_dast = wb.create_sheet("DAST — ZAP")
        make_sheet(ws_dast, zap_rows, "DAST — ZAP")

    # --- Foglio Riepilogo ---
    ws2 = wb.create_sheet("Riepilogo")
    now = datetime.now().strftime("%d/%m/%Y %H:%M")

    summary = [
        ("Report generato il",  now),
        ("", ""),
        ("TOTALE FINDING",       len(rows)),
        ("  CRITICO",            sum(1 for r in rows if r["Severità"] == "CRITICO")),
        ("  ALTO",               sum(1 for r in rows if r["Severità"] == "ALTO")),
        ("  MEDIO",              sum(1 for r in rows if r["Severità"] == "MEDIO")),
        ("  BASSO",              sum(1 for r in rows if r["Severità"] == "BASSO")),
        ("", ""),
        ("SAST — Semgrep",      len(semgrep_rows) if semgrep_rows else 0),
        ("DAST — ZAP",          len(zap_rows)     if zap_rows     else 0),
        ("", ""),
        ("Stato fix",           "Da compilare in aula"),
    ]

    sev_fills = {
        "  CRITICO": COLORS["CRITICO"],
        "  ALTO":    COLORS["ALTO"],
        "  MEDIO":   COLORS["MEDIO"],
        "  BASSO":   COLORS["BASSO"],
    }

    for ri, (label, value) in enumerate(summary, 1):
        ca = ws2.cell(row=ri, column=1, value=label)
        cb = ws2.cell(row=ri, column=2, value=value)
        ca.font = Font(bold=bool(label and not label.startswith(" ")), size=10)
        cb.font = Font(size=10)
        if label in sev_fills:
            for c in (ca, cb):
                c.fill = PatternFill("solid", fgColor=sev_fills[label])

    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 28

    wb.save(output_path)
    n_crit = sum(1 for r in rows if r["Severità"] == "CRITICO")
    print(f"✅ Excel salvato: {output_path}  "
          f"({len(rows)} finding totali, {n_crit} critici, "
          f"{len(wb.sheetnames)} fogli)")


# ---------------------------------------------------------------------------
# SEMGREP — RUN INTERNO
# ---------------------------------------------------------------------------

def find_semgrep():
    s = shutil.which("semgrep")
    if s:
        return s
    import glob
    for p in [
        os.path.expanduser(r"~\AppData\Local\Programs\Python\Python3*\Scripts\semgrep.exe"),
        os.path.expanduser(r"~\AppData\Roaming\Python\Python3*\Scripts\semgrep.exe"),
    ]:
        m = glob.glob(p)
        if m:
            return m[0]
    return None


def run_semgrep(target, config):
    semgrep_bin = find_semgrep()
    if not semgrep_bin:
        print("❌ semgrep non trovato. Installa con: pip install semgrep")
        sys.exit(1)

    if not os.path.isfile(config):
        candidate = os.path.join(os.path.dirname(os.path.abspath(__file__)), config)
        if os.path.isfile(candidate):
            config = candidate
        else:
            print(f"❌ Config non trovata: {config}")
            sys.exit(1)

    if not os.path.exists(target):
        print(f"❌ Target non trovato: {target}")
        sys.exit(1)

    print(f"🔍 Scansione SAST: {target}  (config: {config})")
    result = subprocess.run(
        [semgrep_bin, "--config", config, target, "--json"],
        capture_output=True, text=True,
    )
    if result.returncode not in (0, 1):
        print(f"❌ Semgrep errore (exit {result.returncode}):\n{result.stderr[-600:]}")
        sys.exit(1)
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError:
        print("❌ Output semgrep non è JSON valido.")
        sys.exit(1)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="SAST (Semgrep) + DAST (ZAP) → Excel/CSV unificato",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Esempi:
  python3 semgrep2excel.py --json semgrep.json
  python3 semgrep2excel.py --zap zap_report.json
  python3 semgrep2excel.py --zap zap_report.xml
  python3 semgrep2excel.py --json semgrep.json --zap zap_report.json
  python3 semgrep2excel.py --json semgrep.json --zap zap_report.xml -o report_finale.xlsx
  python3 semgrep2excel.py UserService.java --zap zap_report.json
        """
    )
    parser.add_argument("target",    nargs="?",    default=None,
                        help="File/directory da scansionare con Semgrep")
    parser.add_argument("--json","-j", metavar="FILE.json", default=None,
                        help="JSON già prodotto da Semgrep")
    parser.add_argument("--zap","-z",  metavar="FILE",      default=None,
                        help="Report ZAP in JSON o XML")
    parser.add_argument("--config","-c", default="semgrep-rules-java.yaml",
                        help="Regole Semgrep (default: semgrep-rules-java.yaml)")
    parser.add_argument("--csv",     action="store_true",
                        help="Genera CSV invece di Excel")
    parser.add_argument("-o","--output", default=None,
                        help="Nome file output")
    args = parser.parse_args()

    if not args.json and not args.target and not args.zap and sys.stdin.isatty():
        parser.print_help()
        print("\n❌ Specifica almeno --json, --zap, o un target.")
        sys.exit(1)

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M")
    ext         = "csv" if args.csv else "xlsx"
    output_path = args.output or f"report_sicurezza_{timestamp}.{ext}"

    semgrep_rows, zap_rows = [], []

    # --- Carica finding Semgrep ---
    if args.json:
        if not os.path.isfile(args.json):
            print(f"❌ File JSON non trovato: {args.json}")
            sys.exit(1)
        print(f"📂 Lettura Semgrep JSON: {args.json}")
        with open(args.json, encoding="utf-8") as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError as e:
                print(f"❌ JSON non valido: {e}")
                sys.exit(1)
        semgrep_rows = parse_semgrep(data)
        print(f"   → {len(semgrep_rows)} finding SAST")

    elif args.target:
        data = run_semgrep(args.target, args.config)
        semgrep_rows = parse_semgrep(data)
        print(f"   → {len(semgrep_rows)} finding SAST")

    elif not sys.stdin.isatty() and not args.zap:
        print("📥 Lettura Semgrep da stdin…")
        try:
            data = json.load(sys.stdin)
            semgrep_rows = parse_semgrep(data)
            print(f"   → {len(semgrep_rows)} finding SAST")
        except json.JSONDecodeError as e:
            print(f"❌ JSON stdin non valido: {e}")
            sys.exit(1)

    # --- Carica alert ZAP ---
    if args.zap:
        zap_rows = load_zap(args.zap)
        print(f"   → {len(zap_rows)} alert DAST")

    all_rows = sort_rows(semgrep_rows + zap_rows)

    if not all_rows:
        print("⚠  Nessun finding da nessuna fonte.")
        sys.exit(0)

    if args.csv:
        to_csv(all_rows, output_path)
    else:
        to_excel(all_rows, output_path,
                 semgrep_rows=semgrep_rows or None,
                 zap_rows=zap_rows or None)


if __name__ == "__main__":
    main()
