#!/usr/bin/env python3
"""
semgrep2excel.py — Lancia Semgrep e salva i finding in Excel (.xlsx) o CSV
Compatibile con Windows e macOS | Python 3.8+

USO DIRETTO (consigliato — lancia semgrep internamente):
  # Scansione su file singolo:
  python3 semgrep2excel.py UserService.java

  # Scansione su directory intera:
  python3 semgrep2excel.py src/

  # Config personalizzata:
  python3 semgrep2excel.py UserService.java --config semgrep-rules-java.yaml

  # Output CSV invece di Excel:
  python3 semgrep2excel.py UserService.java --csv

  # Nome file output personalizzato:
  python3 semgrep2excel.py UserService.java -o report_gruppoA.xlsx

USO PIPE (compatibilità con versione precedente):
  semgrep --config rules.yaml . --json | python3 semgrep2excel.py

INSTALLAZIONE DIPENDENZE (una sola volta):
  pip install openpyxl
"""

import json
import sys
import csv
import os
import subprocess
import shutil
import argparse
from datetime import datetime

# Mappa severità → priorità di fix (per ordinamento)
SEVERITY_ORDER = {"ERROR": 1, "WARNING": 2, "INFO": 3}

# Mappa CWE → descrizione leggibile
CWE_LABELS = {
    "CWE-89":  "SQL Injection",
    "CWE-79":  "Cross-Site Scripting (XSS)",
    "CWE-78":  "Command Injection",
    "CWE-22":  "Path Traversal",
    "CWE-798": "Credenziali Hardcoded",
    "CWE-614": "Cookie non sicuro (no HttpOnly/Secure)",
    "CWE-209": "Esposizione Stack Trace",
    "CWE-502": "Deserializzazione non sicura",
    "CWE-918": "SSRF",
    "CWE-347": "JWT - Algoritmo non verificato",
    "CWE-532": "Dati sensibili nei log",
    "CWE-312": "Secret in testo chiaro",
}

# Mappa severità → etichetta italiana
SEVERITY_IT = {
    "ERROR":   "CRITICO",
    "WARNING": "ALTO",
    "INFO":    "MEDIO",
}


def parse_findings(data):
    """Trasforma il JSON Semgrep in una lista di dizionari piatti."""
    rows = []
    for f in data.get("results", []):
        meta     = f["extra"].get("metadata", {})
        def to_str(val):
            """Normalizza stringa o lista in stringa singola."""
            if isinstance(val, list):
                return ", ".join(str(v) for v in val) if val else ""
            return val or ""

        cwe_raw  = to_str(meta.get("cwe", ""))
        cwe_id   = cwe_raw.split(",")[0].strip() if cwe_raw.startswith("CWE-") else ""
        cwe_desc = CWE_LABELS.get(cwe_id, "")
        owasp    = to_str(meta.get("owasp", ""))
        sev_raw  = f["extra"].get("severity", "INFO")

        rows.append({
            "File":           f.get("path", ""),
            "Riga":           f["start"]["line"],
            "Regola":         f.get("check_id", ""),
            "Severità":       SEVERITY_IT.get(sev_raw, sev_raw),
            "CWE":            cwe_id,
            "Descrizione CWE":cwe_desc,
            "OWASP":          owasp,
            "Messaggio":      f["extra"].get("message", ""),
            "Stato fix":      "Aperto",      # campo compilabile in aula
            "Note":           "",            # campo compilabile in aula
            "_sev_order":     SEVERITY_ORDER.get(sev_raw, 9),
        })

    # Ordina: prima per severità, poi per file, poi per riga
    rows.sort(key=lambda r: (r["_sev_order"], r["File"], r["Riga"]))
    for r in rows:
        del r["_sev_order"]

    return rows


def to_csv(rows, output_path):
    """Scrive i finding in CSV UTF-8 con BOM (per Excel su Windows)."""
    if not rows:
        print("⚠  Nessun finding trovato. CSV non generato.")
        return

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

    print(f"✅ CSV salvato: {output_path}  ({len(rows)} finding)")


def to_excel(rows, output_path):
    """Scrive i finding in Excel con colori, larghezze e filtri."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("⚠  openpyxl non trovato. Installa con: pip install openpyxl")
        print("   In alternativa usa --csv")
        sys.exit(1)

    if not rows:
        print("⚠  Nessun finding trovato. Excel non generato.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Finding SAST"

    # --- Colori ---
    COL_HEADER  = "1F4E79"   # blu scuro
    COL_CRITICO = "FCE4D6"   # arancio chiaro
    COL_ALTO    = "FFF2CC"   # giallo chiaro
    COL_MEDIO   = "E2EFDA"   # verde chiaro
    COL_WHITE   = "FFFFFF"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Intestazione ---
    headers = list(rows[0].keys())
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = PatternFill("solid", fgColor=COL_HEADER)
        cell.font      = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = border

    ws.row_dimensions[1].height = 28

    # --- Dati ---
    fill_map = {
        "CRITICO": PatternFill("solid", fgColor=COL_CRITICO),
        "ALTO":    PatternFill("solid", fgColor=COL_ALTO),
        "MEDIO":   PatternFill("solid", fgColor=COL_MEDIO),
    }

    for row_idx, row in enumerate(rows, 2):
        sev   = row.get("Severità", "")
        fill  = fill_map.get(sev, PatternFill("solid", fgColor=COL_WHITE))

        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
            cell.fill      = fill
            cell.font      = Font(size=10)
            cell.border    = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # --- Larghezze colonne ---
    col_widths = {
        "File": 30, "Riga": 8, "Regola": 28, "Severità": 12,
        "CWE": 10, "Descrizione CWE": 26, "OWASP": 14,
        "Messaggio": 50, "Stato fix": 14, "Note": 22,
    }
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = \
            col_widths.get(h, 16)

    # --- Filtri automatici ---
    ws.auto_filter.ref = ws.dimensions

    # --- Foglio riepilogo ---
    ws2 = wb.create_sheet("Riepilogo")
    ws2["A1"] = "Report generato il"
    ws2["B1"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws2["A2"] = "Tool"
    ws2["B2"] = "Semgrep OSS"
    ws2["A3"] = "Finding totali"
    ws2["B3"] = len(rows)
    ws2["A4"] = "CRITICO"
    ws2["B4"] = sum(1 for r in rows if r["Severità"] == "CRITICO")
    ws2["A5"] = "ALTO"
    ws2["B5"] = sum(1 for r in rows if r["Severità"] == "ALTO")
    ws2["A6"] = "MEDIO"
    ws2["B6"] = sum(1 for r in rows if r["Severità"] == "MEDIO")
    ws2["A8"] = "Stato fix"
    ws2["B8"] = "Da compilare in aula"

    for row in ws2.iter_rows(min_row=1, max_row=8, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True, size=10)

    wb.save(output_path)
    print(f"✅ Excel salvato: {output_path}  ({len(rows)} finding, "
          f"{sum(1 for r in rows if r['Severità']=='CRITICO')} critici)")


def find_semgrep():
    """Trova l'eseguibile semgrep su Windows e macOS/Linux."""
    # Cerca prima nel PATH standard
    semgrep = shutil.which("semgrep")
    if semgrep:
        return semgrep

    # Percorsi comuni su Windows (pip install --user)
    win_paths = [
        os.path.expanduser(r"~\AppData\Local\Programs\Python\Python3*\Scripts\semgrep.exe"),
        os.path.expanduser(r"~\AppData\Roaming\Python\Python3*\Scripts\semgrep.exe"),
    ]
    for p in win_paths:
        import glob
        matches = glob.glob(p)
        if matches:
            return matches[0]

    return None


def run_semgrep(target, config):
    """Lancia semgrep sul target e restituisce il JSON dei risultati."""
    semgrep_bin = find_semgrep()

    if not semgrep_bin:
        print("❌ semgrep non trovato nel PATH.")
        print("   Installa con: pip install semgrep")
        sys.exit(1)

    # Cerca il file di config nella stessa cartella dello script se non specificato
    if not os.path.isfile(config):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        candidate  = os.path.join(script_dir, config)
        if os.path.isfile(candidate):
            config = candidate
        else:
            print(f"❌ Config non trovata: {config}")
            print(f"   Metti '{config}' nella stessa cartella di semgrep2excel.py")
            sys.exit(1)

    if not os.path.exists(target):
        print(f"❌ Target non trovato: {target}")
        sys.exit(1)

    print(f"🔍 Scansione in corso: {target}  (config: {config})")

    result = subprocess.run(
        [semgrep_bin, "--config", config, target, "--json"],
        capture_output=True,
        text=True,
    )

    # semgrep esce con codice 1 se trova finding — è normale
    if result.returncode not in (0, 1):
        print(f"❌ Semgrep ha restituito un errore (exit {result.returncode}):")
        print(result.stderr[-800:] if result.stderr else "(nessun dettaglio)")
        sys.exit(1)

    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError:
        print("❌ Output semgrep non è JSON valido.")
        print("Stderr:", result.stderr[-400:])
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(
        description="Converte output Semgrep in Excel o CSV",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Esempi:
  python3 semgrep2excel.py --json output.json
  python3 semgrep2excel.py --json output.json -o report_gruppoA.xlsx
  python3 semgrep2excel.py --json output.json --csv
  python3 semgrep2excel.py UserService.java
  python3 semgrep2excel.py src/
  semgrep --config rules.yaml . --json | python3 semgrep2excel.py
        """
    )
    parser.add_argument(
        "target",
        nargs="?",
        default=None,
        help="File o directory da scansionare (es. UserService.java oppure src/)"
    )
    parser.add_argument(
        "--json", "-j",
        metavar="FILE.json",
        default=None,
        help="File JSON già prodotto da semgrep (es. --json output.json)"
    )
    parser.add_argument(
        "--config", "-c",
        default="semgrep-rules-java.yaml",
        help="File di regole Semgrep (default: semgrep-rules-java.yaml)"
    )
    parser.add_argument(
        "--csv",
        action="store_true",
        help="Genera CSV invece di Excel"
    )
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="Nome file output (default: semgrep_YYYYMMDD_HHMM.xlsx/.csv)"
    )
    args = parser.parse_args()

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M")
    ext         = "csv" if args.csv else "xlsx"
    output_path = args.output or f"semgrep_{timestamp}.{ext}"

    # --- MODALITÀ 1: file JSON già pronto ---
    if args.json:
        if not os.path.isfile(args.json):
            print(f"❌ File JSON non trovato: {args.json}")
            sys.exit(1)
        print(f"📂 Lettura JSON: {args.json}")
        with open(args.json, encoding="utf-8") as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError as e:
                print(f"❌ JSON non valido: {e}")
                sys.exit(1)

    # --- MODALITÀ 2: target passato come argomento (lancia semgrep) ---
    elif args.target:
        data = run_semgrep(args.target, args.config)

    # --- MODALITÀ 3: pipe da semgrep (retrocompatibilità) ---
    elif not sys.stdin.isatty():
        print("📥 Lettura da stdin (modalità pipe)...")
        try:
            data = json.load(sys.stdin)
        except json.JSONDecodeError as e:
            print(f"❌ JSON non valido da stdin: {e}")
            sys.exit(1)

    # --- Nessun input ---
    else:
        parser.print_help()
        print("\n❌ Specifica --json FILE.json, un target, oppure usa la modalità pipe.")
        sys.exit(1)

    rows = parse_findings(data)

    if args.csv:
        to_csv(rows, output_path)
    else:
        to_excel(rows, output_path)


if __name__ == "__main__":
    main()
